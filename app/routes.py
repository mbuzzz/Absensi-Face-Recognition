# app/routes.py

from flask import (Blueprint, render_template, redirect, url_for, flash, 
                   Response, request, jsonify, current_app, send_file)
from flask_login import login_user, logout_user, current_user, login_required
from functools import wraps
from sqlalchemy.exc import OperationalError
from werkzeug.utils import secure_filename
import cv2
import face_recognition
import numpy as np
import os
import requests
from datetime import datetime, time as dt_time, timedelta
import time
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app import db
from app.models import User, Siswa, Kelas, Absensi, Pengaturan
from app.forms import (LoginForm, SiswaForm, SiswaEditForm, KelasForm, 
                       UserForm, UserEditForm, LaporanFilterForm, 
                       KehadiranHarianFilterForm, KehadiranUpdateForm, PengaturanForm)
from app.face_logic import train_faces, load_known_faces

bp = Blueprint('main', __name__)

# --- Decorator untuk Hak Akses ---
def role_required(roles):
    """Decorator untuk membatasi akses berdasarkan peran pengguna."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash('Anda tidak memiliki hak akses untuk halaman ini.', 'danger')
                return redirect(url_for('main.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- State, Cache, dan Fungsi Absensi Otomatis ---
auto_absen_aktif = False
known_face_encodings, known_face_ids = [], []
siswa_data_cache = {}

def refresh_data():
    """Memuat ulang data wajah dan cache data siswa."""
    global known_face_encodings, known_face_ids, siswa_data_cache
    print("Memuat data wajah dan cache siswa...")
    known_face_encodings, known_face_ids = load_known_faces()
    try:
        siswas = Siswa.query.join(Kelas).with_entities(Siswa.id, Siswa.nama, Kelas.nama_kelas).all()
        siswa_data_cache = {s_id: (nama, kelas) for s_id, nama, kelas in siswas}
        print("Data wajah dan cache siswa telah berhasil dimuat.")
    except OperationalError:
        print("Peringatan: Gagal memuat cache siswa karena tabel belum ada.")
        siswa_data_cache = {}
    except Exception as e:
        print(f"Error tidak terduga saat memuat cache siswa: {e}")
        siswa_data_cache = {}

def get_setting(key, default=None, is_time=False):
    """Mengambil nilai pengaturan dari database."""
    setting = Pengaturan.query.filter_by(key=key).first()
    if setting and setting.value:
        if is_time:
            try:
                return datetime.strptime(setting.value, '%H:%M').time()
            except (ValueError, TypeError): return default
        return setting.value
    return default

def set_setting(key, value):
    """Menyimpan nilai pengaturan ke database."""
    setting = Pengaturan.query.filter_by(key=key).first()
    if isinstance(value, (datetime, dt_time)):
        value = value.strftime('%H:%M')
    if setting:
        setting.value = str(value)
    else:
        setting = Pengaturan(key=key, value=str(value))
        db.session.add(setting)
    db.session.commit()

def catat_absensi(siswa_id):
    """Mencatat absensi dan mengembalikan status apakah catatan baru dibuat."""
    with current_app.app_context():
        today = datetime.utcnow().date()
        absen_hari_ini = Absensi.query.filter(Absensi.siswa_id == siswa_id, db.func.date(Absensi.timestamp) == today).first()
        if not absen_hari_ini:
            now_time = datetime.now().time()
            batas_terlambat = get_setting('waktu_terlambat', default=dt_time(7, 0), is_time=True)
            status = 'Hadir' if now_time <= batas_terlambat else 'Terlambat'
            absen = Absensi(siswa_id=siswa_id, status=status, timestamp=datetime.utcnow())
            db.session.add(absen); db.session.commit()
            kirim_notifikasi_n8n(absen)
            print(f"Absensi '{status}' untuk siswa ID {siswa_id} tercatat.")
            return True
    return False

def kirim_notifikasi_n8n(absen_obj):
    """Mengirim notifikasi ke n8n melalui webhook dengan log yang lebih baik."""
    webhook_url = current_app.config.get('N8N_WEBHOOK_URL')
    if not webhook_url:
        print("Peringatan: N8N_WEBHOOK_URL tidak diatur. Notifikasi tidak dikirim.")
        return

    siswa = absen_obj.siswa
    if not siswa:
        print(f"Error: Tidak dapat menemukan data siswa untuk Absensi ID: {absen_obj.id}")
        return

    payload = {
        'nis': siswa.nis, 'nama': siswa.nama, 'kelas': siswa.kelas.nama_kelas,
        'nomor_orang_tua': siswa.nomor_orang_tua, 'status': absen_obj.status,
        'keterangan': absen_obj.keterangan,
        'waktu_absen_utc': absen_obj.timestamp.isoformat() + 'Z',
        'waktu_absen_lokal': (absen_obj.timestamp + timedelta(hours=7)).strftime('%d-%m-%Y %H:%M:%S')
    }
    try:
        print(f"Mengirim data ke n8n: {payload}")
        response = requests.post(webhook_url, json=payload, timeout=10)
        print(f"n8n response status: {response.status_code}")
        print(f"n8n response body: {response.text}")
        if response.status_code == 200:
            print(f"Notifikasi untuk {siswa.nama} ({absen_obj.status}) berhasil dikirim ke n8n.")
        else:
            print(f"Gagal mengirim notifikasi ke n8n. Status: {response.status_code}")
    except requests.exceptions.RequestException as e:
        print(f"Gagal mengirim notifikasi ke n8n (Request Exception): {e}")

# --- Helper untuk Pengaturan & Context Processor ---
@bp.context_processor
def inject_settings():
    """Menyediakan variabel pengaturan ke semua template."""
    with current_app.app_context():
        nama_aplikasi = get_setting('nama_aplikasi', 'Absensi Wajah')
        logo_path = get_setting('logo_aplikasi')
    return dict(nama_aplikasi=nama_aplikasi, logo_path=logo_path)

# --- Rute Utama dan Kontrol Kamera ---
@bp.route('/')
def index():
    return render_template('index.html', title='Live Absensi')

@bp.route('/start_absen', methods=['POST'])
@login_required
@role_required(['Admin', 'Petugas'])
def start_absen():
    global auto_absen_aktif; auto_absen_aktif = True; return jsonify({'status': 'sukses'})

@bp.route('/stop_absen', methods=['POST'])
@login_required
@role_required(['Admin', 'Petugas'])
def stop_absen():
    global auto_absen_aktif; auto_absen_aktif = False; return jsonify({'status': 'sukses'})

@bp.route('/absen_status')
def absen_status():
    return jsonify({'aktif': auto_absen_aktif})

def gen_frames():
    """Generator function untuk video streaming dengan umpan balik visual."""
    video_capture = cv2.VideoCapture(0)
    if not video_capture.isOpened():
        print("Error: Tidak bisa membuka kamera."); return

    last_face_locations, last_face_names, last_face_ids = [], [], []
    frame_count = 0
    PROCESS_EVERY_N_FRAMES = 5
    RESIZE_FACTOR = 0.25
    recently_recorded = {} # Menyimpan ID siswa yang baru saja tercatat

    try:
        while True:
            try:
                success, frame = video_capture.read()
                if not success: break

                if frame_count % PROCESS_EVERY_N_FRAMES == 0:
                    small_frame = cv2.resize(frame, (0, 0), fx=RESIZE_FACTOR, fy=RESIZE_FACTOR)
                    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                    
                    current_face_locations = face_recognition.face_locations(rgb_small_frame)
                    current_face_encodings = face_recognition.face_encodings(rgb_small_frame, current_face_locations)
                    
                    current_face_names, current_face_ids = [], []
                    for face_encoding in current_face_encodings:
                        matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.5)
                        name, kelas_info, siswa_id = "Tidak Dikenal", "", None

                        if True in matches:
                            first_match_index = matches.index(True)
                            siswa_id = known_face_ids[first_match_index]
                            name, kelas_info = siswa_data_cache.get(siswa_id, ("Siswa Dihapus", ""))
                            if auto_absen_aktif:
                                if catat_absensi(siswa_id):
                                    recently_recorded[siswa_id] = time.time() # Tandai waktu pencatatan
                        
                        current_face_names.append(f"{name} ({kelas_info})")
                        current_face_ids.append(siswa_id)
                    
                    last_face_locations, last_face_names, last_face_ids = current_face_locations, current_face_names, current_face_ids
                
                frame_count += 1

                for (top, right, bottom, left), name, siswa_id in zip(last_face_locations, last_face_names, last_face_ids):
                    top = int(top / RESIZE_FACTOR)
                    right = int(right / RESIZE_FACTOR)
                    bottom = int(bottom / RESIZE_FACTOR)
                    left = int(left / RESIZE_FACTOR)
                    
                    color = (255, 0, 0) # Merah untuk tidak dikenal
                    if siswa_id in recently_recorded and time.time() - recently_recorded[siswa_id] < 3:
                        color = (0, 255, 0) # Hijau terang untuk berhasil tercatat
                    elif siswa_id is not None:
                        color = (0, 180, 216) # Biru untuk sudah dikenali/sudah absen

                    cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                    cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                    cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.7, (255, 255, 255), 1)

                ret, buffer = cv2.imencode('.jpg', frame)
                yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            
            except Exception as e:
                print(f"Error di dalam loop gen_frames: {e}"); continue
    finally:
        print("Melepaskan kamera."); video_capture.release()

@bp.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

# --- Rute Otentikasi & Dashboard ---
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('main.dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Username atau password tidak valid', 'danger'); return redirect(url_for('main.login'))
        login_user(user, remember=form.remember_me.data)
        return redirect(url_for('main.dashboard'))
    return render_template('auth/login.html', title='Login', form=form)

@bp.route('/logout')
def logout():
    logout_user(); return redirect(url_for('main.index'))

@bp.route('/dashboard')
@login_required
def dashboard():
    return render_template('admin/dashboard.html', title='Dashboard')

@bp.route('/dashboard_data')
@login_required
def dashboard_data():
    today_start = datetime.combine(datetime.utcnow().date(), dt_time.min)
    today_end = datetime.combine(datetime.utcnow().date(), dt_time.max)
    query_today = Absensi.query.filter(Absensi.timestamp.between(today_start, today_end))
    total_siswa = Siswa.query.count()
    rekap_harian = {'Hadir': 0, 'Terlambat': 0, 'Izin': 0, 'Sakit': 0, 'Alpha': 0, 'total_siswa': total_siswa}
    for absen in query_today.all():
        if absen.status in rekap_harian: rekap_harian[absen.status] += 1
    labels, tren_hadir, tren_terlambat, tren_izin, tren_sakit, tren_alpha = [], [], [], [], [], []
    for i in range(6, -1, -1):
        hari = datetime.utcnow().date() - timedelta(days=i)
        labels.append(hari.strftime('%a, %d'))
        start, end = datetime.combine(hari, dt_time.min), datetime.combine(hari, dt_time.max)
        tren_hadir.append(Absensi.query.filter(Absensi.timestamp.between(start, end), Absensi.status == 'Hadir').count())
        tren_terlambat.append(Absensi.query.filter(Absensi.timestamp.between(start, end), Absensi.status == 'Terlambat').count())
        tren_izin.append(Absensi.query.filter(Absensi.timestamp.between(start, end), Absensi.status == 'Izin').count())
        tren_sakit.append(Absensi.query.filter(Absensi.timestamp.between(start, end), Absensi.status == 'Sakit').count())
        tren_alpha.append(Absensi.query.filter(Absensi.timestamp.between(start, end), Absensi.status == 'Alpha').count())
    tren_mingguan = {'labels': labels, 'hadir': tren_hadir, 'terlambat': tren_terlambat, 'izin': tren_izin, 'sakit': tren_sakit, 'alpha': tren_alpha}
    return jsonify({'rekap_harian': rekap_harian, 'tren_mingguan': tren_mingguan})

# --- Rute Pengaturan ---
@bp.route('/admin/pengaturan', methods=['GET', 'POST'])
@login_required
@role_required(['Admin'])
def pengaturan():
    form = PengaturanForm()
    if form.validate_on_submit():
        set_setting('nama_aplikasi', form.nama_aplikasi.data)
        set_setting('nama_sekolah', form.nama_sekolah.data)
        set_setting('waktu_terlambat', form.waktu_terlambat.data)
        set_setting('waktu_alpha', form.waktu_alpha.data)
        if form.logo_aplikasi.data:
            f = form.logo_aplikasi.data
            filename = secure_filename(f.filename)
            logo_filename = f"logo.{filename.rsplit('.', 1)[1].lower()}"
            f.save(os.path.join(current_app.static_folder, logo_filename))
            set_setting('logo_aplikasi', logo_filename)
        flash('Pengaturan berhasil disimpan!', 'success'); return redirect(url_for('main.pengaturan'))
    form.nama_aplikasi.data = get_setting('nama_aplikasi', 'Absensi Wajah')
    form.nama_sekolah.data = get_setting('nama_sekolah', 'Nama Sekolah Anda')
    form.waktu_terlambat.data = get_setting('waktu_terlambat', dt_time(7, 0), is_time=True)
    form.waktu_alpha.data = get_setting('waktu_alpha', dt_time(9, 0), is_time=True)
    return render_template('admin/pengaturan.html', title='Pengaturan Aplikasi', form=form)

# --- CRUD SISWA ---
@bp.route('/admin/siswa')
@login_required
@role_required(['Admin', 'Wali Kelas', 'BK'])
def crud_siswa():
    if current_user.role == 'Wali Kelas':
        if not current_user.kelas_id:
            flash('Akun Anda belum ditautkan ke kelas manapun.', 'warning'); return render_template('admin/crud_siswa.html', title='Manajemen Siswa', siswas=[])
        siswas = Siswa.query.filter_by(kelas_id=current_user.kelas_id).order_by(Siswa.nama).all()
    else:
        siswas = Siswa.query.order_by(Siswa.nama).all()
    return render_template('admin/crud_siswa.html', title='Manajemen Siswa', siswas=siswas)

@bp.route('/admin/siswa/tambah', methods=['GET', 'POST'])
@login_required
@role_required(['Admin', 'BK'])
def tambah_siswa():
    form = SiswaForm()
    if form.validate_on_submit():
        f = form.foto.data
        unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(f.filename)}"
        f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename))
        siswa = Siswa(nis=form.nis.data, nama=form.nama.data, nomor_orang_tua=form.nomor_orang_tua.data, kelas_id=form.kelas.data.id, foto_path=unique_filename)
        db.session.add(siswa); db.session.commit(); train_faces(); refresh_data()
        flash('Data siswa berhasil ditambahkan!', 'success'); return redirect(url_for('main.crud_siswa'))
    return render_template('admin/tambah_edit_siswa.html', title='Tambah Siswa', form=form)

@bp.route('/admin/siswa/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['Admin', 'Wali Kelas', 'BK'])
def edit_siswa(id):
    siswa = Siswa.query.get_or_404(id)
    if current_user.role == 'Wali Kelas' and siswa.kelas_id != current_user.kelas_id:
        flash('Anda tidak berhak mengedit data siswa dari kelas lain.', 'danger'); return redirect(url_for('main.crud_siswa'))
    form = SiswaEditForm(obj=siswa)
    if form.validate_on_submit():
        if form.foto.data:
            if siswa.foto_path and os.path.exists(os.path.join(current_app.config['UPLOAD_FOLDER'], siswa.foto_path)):
                os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], siswa.foto_path))
            f = form.foto.data
            unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(f.filename)}"
            f.save(os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename))
            siswa.foto_path = unique_filename
        siswa.nis = form.nis.data; siswa.nama = form.nama.data; siswa.nomor_orang_tua = form.nomor_orang_tua.data; siswa.kelas_id = form.kelas.data.id
        db.session.commit(); train_faces(); refresh_data()
        flash('Data siswa berhasil diperbarui!', 'success'); return redirect(url_for('main.crud_siswa'))
    return render_template('admin/tambah_edit_siswa.html', title='Edit Siswa', form=form, siswa=siswa)

@bp.route('/admin/siswa/hapus/<int:id>', methods=['POST'])
@login_required
@role_required(['Admin'])
def hapus_siswa(id):
    siswa = Siswa.query.get_or_404(id)
    if siswa.foto_path and os.path.exists(os.path.join(current_app.config['UPLOAD_FOLDER'], siswa.foto_path)):
        os.remove(os.path.join(current_app.config['UPLOAD_FOLDER'], siswa.foto_path))
    Absensi.query.filter_by(siswa_id=id).delete()
    db.session.delete(siswa); db.session.commit(); train_faces(); refresh_data()
    flash('Data siswa dan semua riwayat absensinya berhasil dihapus.', 'success'); return redirect(url_for('main.crud_siswa'))

# --- CRUD KELAS ---
@bp.route('/admin/kelas', methods=['GET', 'POST'])
@login_required
@role_required(['Admin'])
def crud_kelas():
    form = KelasForm()
    if form.validate_on_submit():
        kelas = Kelas(nama_kelas=form.nama_kelas.data, wali_kelas_nama=form.wali_kelas_nama.data)
        db.session.add(kelas); db.session.commit()
        flash('Kelas baru berhasil ditambahkan!', 'success'); return redirect(url_for('main.crud_kelas'))
    kelasList = Kelas.query.order_by(Kelas.nama_kelas).all()
    return render_template('admin/crud_kelas.html', title='Manajemen Kelas', form=form, kelasList=kelasList)

@bp.route('/admin/kelas/hapus/<int:id>', methods=['POST'])
@login_required
@role_required(['Admin'])
def hapus_kelas(id):
    kelas = Kelas.query.get_or_404(id)
    if kelas.siswas.count() > 0:
        flash('Tidak dapat menghapus kelas karena masih ada siswa di dalamnya.', 'danger'); return redirect(url_for('main.crud_kelas'))
    User.query.filter_by(kelas_id=id).update({User.kelas_id: None})
    db.session.delete(kelas); db.session.commit()
    flash('Kelas berhasil dihapus.', 'success'); return redirect(url_for('main.crud_kelas'))

# --- CRUD USER ---
@bp.route('/admin/user')
@login_required
@role_required(['Admin'])
def crud_user():
    users = User.query.all()
    return render_template('admin/crud_user.html', title='Manajemen Pengguna', users=users)

@bp.route('/admin/user/tambah', methods=['GET', 'POST'])
@login_required
@role_required(['Admin'])
def tambah_user():
    form = UserForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data, role=form.role.data)
        user.set_password(form.password.data)
        if form.role.data == 'Wali Kelas' and form.kelas_asuhan.data:
            user.kelas_id = form.kelas_asuhan.data.id
        db.session.add(user); db.session.commit()
        flash('Pengguna baru berhasil ditambahkan!', 'success'); return redirect(url_for('main.crud_user'))
    return render_template('admin/tambah_edit_user.html', title='Tambah Pengguna', form=form)

@bp.route('/admin/user/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['Admin'])
def edit_user(id):
    user = User.query.get_or_404(id)
    form = UserEditForm(obj=user)
    if form.validate_on_submit():
        user.username = form.username.data; user.email = form.email.data; user.role = form.role.data
        if form.role.data == 'Wali Kelas' and form.kelas_asuhan.data:
            user.kelas_id = form.kelas_asuhan.data.id
        else:
            user.kelas_id = None
        if form.password.data:
            user.set_password(form.password.data)
        db.session.commit()
        flash('Data pengguna berhasil diperbarui!', 'success'); return redirect(url_for('main.crud_user'))
    return render_template('admin/tambah_edit_user.html', title='Edit Pengguna', form=form, user=user)

@bp.route('/admin/user/hapus/<int:id>', methods=['POST'])
@login_required
@role_required(['Admin'])
def hapus_user(id):
    if id == current_user.id:
        flash('Anda tidak dapat menghapus akun Anda sendiri.', 'danger'); return redirect(url_for('main.crud_user'))
    user = User.query.get_or_404(id)
    db.session.delete(user); db.session.commit()
    flash('Pengguna berhasil dihapus.', 'success'); return redirect(url_for('main.crud_user'))

# --- LAPORAN & EDIT KEHADIRAN ---
@bp.route('/admin/laporan_log', methods=['GET', 'POST'])
@login_required
@role_required(['Admin', 'Wali Kelas', 'BK'])
def laporan_log():
    form = LaporanFilterForm()
    rekap = {'Hadir': 0, 'Terlambat': 0, 'Izin': 0, 'Sakit': 0, 'Alpha': 0}
    laporan_data = []

    if form.validate_on_submit():
        start_datetime = datetime.combine(form.tanggal_mulai.data, dt_time.min)
        end_datetime = datetime.combine(form.tanggal_akhir.data, dt_time.max)
        
        query = Absensi.query.join(Siswa).filter(Absensi.timestamp.between(start_datetime, end_datetime))
        if form.kelas.data:
            query = query.filter(Siswa.kelas_id == form.kelas.data.id)
        elif current_user.role == 'Wali Kelas' and current_user.kelas_id:
            query = query.filter(Siswa.kelas_id == current_user.kelas_id)
        
        laporan_data = query.order_by(Absensi.timestamp.desc()).all()
        for absen in laporan_data:
            if absen.status in rekap: rekap[absen.status] += 1

        if form.export.data:
            return export_laporan_to_excel(laporan_data, rekap, form.tanggal_mulai.data, form.tanggal_akhir.data, form.kelas.data)
            
    return render_template('admin/laporan_log.html', title='Log Kehadiran', form=form, laporan=laporan_data, rekap=rekap)

@bp.route('/admin/kehadiran_harian', methods=['GET', 'POST'])
@login_required
@role_required(['Admin', 'Wali Kelas', 'BK'])
def edit_kehadiran_harian():
    form = KehadiranHarianFilterForm()
    update_form = KehadiranUpdateForm()
    rekap = {'Hadir': 0, 'Terlambat': 0, 'Izin': 0, 'Sakit': 0, 'Alpha': 0}
    
    tanggal_laporan = datetime.utcnow().date()
    if form.validate_on_submit():
        tanggal_laporan = form.tanggal.data
    elif 'tanggal' in request.args:
        try:
            tanggal_laporan = datetime.strptime(request.args.get('tanggal'), '%Y-%m-%d').date()
        except (ValueError, TypeError): pass
    
    form.tanggal.data = tanggal_laporan

    if current_user.role == 'Wali Kelas':
        if not current_user.kelas_id:
            flash('Akun Anda belum ditautkan ke kelas manapun.', 'warning'); return render_template('admin/kehadiran_harian.html', title='Edit Kehadiran Harian', form=form, update_form=update_form, laporan=[], rekap=rekap, tanggal=tanggal_laporan)
        query_siswa = Siswa.query.filter_by(kelas_id=current_user.kelas_id)
        form.kelas.data = Kelas.query.get(current_user.kelas_id)
    elif form.kelas.data:
        query_siswa = Siswa.query.filter_by(kelas_id=form.kelas.data.id)
    else:
        query_siswa = Siswa.query

    semua_siswa = query_siswa.order_by(Siswa.nama).all()
    id_siswa = [s.id for s in semua_siswa]
    start_datetime = datetime.combine(tanggal_laporan, dt_time.min)
    end_datetime = datetime.combine(tanggal_laporan, dt_time.max)
    absensi_tercatat = {a.siswa_id: a for a in Absensi.query.filter(Absensi.siswa_id.in_(id_siswa), Absensi.timestamp.between(start_datetime, end_datetime)).all()}

    batas_alpha = get_setting('waktu_alpha', default=dt_time(9,0), is_time=True)
    if tanggal_laporan == datetime.utcnow().date() and datetime.now().time() > batas_alpha:
        siswa_belum_absen = [s for s in semua_siswa if s.id not in absensi_tercatat]
        for siswa in siswa_belum_absen:
            alpha_record = Absensi(siswa_id=siswa.id, status='Alpha', keterangan='Dibuat otomatis oleh sistem', timestamp=datetime.combine(tanggal_laporan, batas_alpha), siswa=siswa)
            db.session.add(alpha_record); kirim_notifikasi_n8n(alpha_record)
        if siswa_belum_absen:
            db.session.commit()
            absensi_tercatat = {a.siswa_id: a for a in Absensi.query.filter(Absensi.siswa_id.in_(id_siswa), Absensi.timestamp.between(start_datetime, end_datetime)).all()}

    laporan_final = []
    for siswa in semua_siswa:
        absen = absensi_tercatat.get(siswa.id)
        if absen:
            laporan_final.append(absen); rekap[absen.status] += 1
        else:
            dummy_absen = Absensi(siswa_id=siswa.id, status='Alpha', siswa=siswa); laporan_final.append(dummy_absen); rekap['Alpha'] += 1
    
    return render_template('admin/kehadiran_harian.html', title='Edit Kehadiran Harian', form=form, update_form=update_form, laporan=laporan_final, rekap=rekap, tanggal=tanggal_laporan)

@bp.route('/admin/kehadiran/update/<int:siswa_id>', methods=['POST'])
@login_required
@role_required(['Admin', 'Wali Kelas', 'BK'])
def update_kehadiran(siswa_id):
    form = KehadiranUpdateForm()
    tanggal_str = request.form.get('tanggal')
    
    if form.validate_on_submit():
        tanggal = datetime.strptime(tanggal_str, '%Y-%m-%d').date()
        start_datetime = datetime.combine(tanggal, dt_time.min)
        end_datetime = datetime.combine(tanggal, dt_time.max)
        
        absen = Absensi.query.filter(Absensi.siswa_id == siswa_id, Absensi.timestamp.between(start_datetime, end_datetime)).first()
        if absen:
            absen.status = form.status.data; absen.keterangan = form.keterangan.data
        else:
            absen = Absensi(siswa_id=siswa_id, status=form.status.data, keterangan=form.keterangan.data, timestamp=start_datetime)
            db.session.add(absen)
        db.session.commit()
        kirim_notifikasi_n8n(absen)
        flash(f'Kehadiran siswa berhasil diupdate.', 'success')
    else:
        flash('Gagal mengupdate kehadiran.', 'danger')
    return redirect(url_for('main.edit_kehadiran_harian', tanggal=tanggal_str, kelas_id=request.form.get('kelas_id', '')))

# --- FUNGSI EKSPOR EXCEL ---
def export_laporan_to_excel(laporan_data, rekap, tgl_mulai, tgl_akhir, kelas_filter):
    """Fungsi untuk membuat file Excel dari data laporan."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Kehadiran"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    sheet.merge_cells('A1:H1'); cell = sheet['A1']
    cell.value = 'Laporan Absensi Kehadiran Siswa'; cell.font = title_font; cell.alignment = center_align
    periode_str = f"Periode: {tgl_mulai.strftime('%d-%m-%Y')} s/d {tgl_akhir.strftime('%d-%m-%Y')}"
    kelas_str = f"Kelas: {kelas_filter.nama_kelas}" if kelas_filter else "Kelas: Semua"
    sheet.merge_cells('A2:H2'); sheet['A2'].value = f"{periode_str} | {kelas_str}"; sheet['A2'].alignment = center_align
    sheet.row_dimensions[1].height = 20; sheet.row_dimensions[2].height = 18

    sheet.merge_cells('A4:B4'); sheet['A4'].value = 'REKAPITULASI TOTAL'; sheet['A4'].font = header_font
    rekap_headers = ['Hadir', 'Terlambat', 'Izin', 'Sakit', 'Alpha']
    for i, header in enumerate(rekap_headers):
        cell = sheet.cell(row=5, column=i+1); cell.value = header; cell.font = Font(bold=True); cell.alignment = center_align; cell.border = thin_border
        cell_val = sheet.cell(row=6, column=i+1); cell_val.value = rekap.get(header, 0); cell_val.alignment = center_align; cell_val.border = thin_border

    table_start_row = 8
    headers = ["No", "Tanggal", "Waktu", "NIS", "Nama Siswa", "Kelas", "Status", "Keterangan"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=table_start_row, column=col_num); cell.value = header; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
    
    sheet.column_dimensions['A'].width = 5; sheet.column_dimensions['B'].width = 12; sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 15; sheet.column_dimensions['E'].width = 35; sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 10; sheet.column_dimensions['H'].width = 40

    for idx, data in enumerate(laporan_data, 1):
        row_num = table_start_row + idx
        row_data = [idx, data.timestamp.strftime('%d-%m-%Y'), data.timestamp.strftime('%H:%M:%S'), data.siswa.nis, data.siswa.nama, data.siswa.kelas.nama_kelas, data.status, data.keterangan or '']
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num); cell.value = cell_value; cell.border = thin_border
            cell.alignment = left_align if col_num in [5, 8] else center_align

    workbook.save(output); output.seek(0)
    filename = f"laporan_absensi_{tgl_mulai.strftime('%Y%m%d')}_{tgl_akhir.strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
